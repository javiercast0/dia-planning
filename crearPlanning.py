import openpyxl
from tkinter import Tk, filedialog
import datetime

# Funcion para seleccionar archivos.
def seleccionar_archivo(titulo):
    root = Tk()
    root.withdraw()
    archivo = filedialog.askopenfilename(title=titulo, filetypes=[("Excel files", "*.xlsm *.xlsx")])
    root.destroy()
    return archivo

#Funcion para generar el mapeo
def generar_mapeo(hoja, col_busqueda, cols_interes):
    mapa = {}
    for row in hoja.iter_rows(min_row=1, max_col=max(cols_interes), values_only=True):
        clave = row[col_busqueda - 1]
        if clave:
            mapa[clave] = {c: row[c-1] for c in cols_interes}
    return mapa

# Lista de tiendas que se sirven por la tarde, estas tiendas tienen
# la fruta en AxC y la carne en AxB
TIENDAS_TARDE = {7658, 2172, 26054, 2417, 2498, 7345, 2163, 24032, 7401, 7643, 7725, 25010, 26053, 2473}

# 1. Selección de archivos
files = {
    "actual": seleccionar_archivo("Archivo día actual"),
    "ayer": seleccionar_archivo("Archivo día anterior"),
    "hace_2_dias": seleccionar_archivo("Archivo hace 2 días"),
    "hace_3_dias": seleccionar_archivo("Archivo hace 3 días")
}

# 2. Carga de datos
wbs = {k: openpyxl.load_workbook(v, data_only=True)['Resumen'] for k, v in files.items()}

# 3. Crear mapeos
mapa_actual = generar_mapeo(wbs["actual"], 4, [17, 18, 23, 24])
mapa_ayer = generar_mapeo(wbs["ayer"], 4, [21, 22])
mapa_2_dias = generar_mapeo(wbs["hace_2_dias"], 4, [19, 21, 22])
mapa_3_dias = generar_mapeo(wbs["hace_3_dias"], 4, [19])

# 4. Crear archivo final
new_wb = openpyxl.Workbook()
ws_final = new_wb.active

# 5. Obtener lista de tiendas
tiendas = []
for row in wbs["actual"].iter_rows(min_row=7, min_col=4, max_col=4, values_only=True):
    if row[0] is None: break
    tiendas.append(row[0])

# 6. Procesar los datos
for i, tienda in enumerate(tiendas, start=1):
    ws_final.cell(row=i, column=1, value=tienda)
    
    # --- DATOS FIJOS (Día Actual) ---
    if tienda in mapa_actual:
        ws_final.cell(row=i, column=2, value=mapa_actual[tienda][17]) # Area 11
        ws_final.cell(row=i, column=3, value=mapa_actual[tienda][18]) # Area 31
        ws_final.cell(row=i, column=8, value=mapa_actual[tienda][23]) # Area 91
        ws_final.cell(row=i, column=4, value=mapa_actual[tienda][24]) # Area 41
    
    # --- Coger datos ---
    if tienda in TIENDAS_TARDE:
        if tienda in mapa_2_dias:
            ws_final.cell(row=i, column=5, value=mapa_2_dias[tienda][19])
        if tienda in mapa_ayer:
            ws_final.cell(row=i, column=6, value=mapa_ayer[tienda][21])
            ws_final.cell(row=i, column=7, value=mapa_ayer[tienda][22])
    else:
        if tienda in mapa_3_dias:
            ws_final.cell(row=i, column=5, value=mapa_3_dias[tienda][19])
        if tienda in mapa_2_dias:
            ws_final.cell(row=i, column=6, value=mapa_2_dias[tienda][21])
            ws_final.cell(row=i, column=7, value=mapa_2_dias[tienda][22])

# 7. Guardar
fecha_actual = datetime.datetime.now().strftime("%d-%m").lstrip('0')
new_wb.save(f'Volumetría {fecha_actual} (DRIVE).xlsx')
print("Proceso finalizado con éxito.")